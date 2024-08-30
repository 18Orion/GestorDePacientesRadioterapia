import openpyxl
from libs.globalVars import TREATMENT_DICT, GENDER_GUESSER_DICT, DATE_OPTIONS
from time import time
from datetime import date, datetime 
import gender_guesser.detector as gender
from libs.MySQLdb import MySQLdb
from libs.funcs import printProgressBar, inputLogin

def getNUSHA(nusha):
    if nusha:
        if type(nusha)==int:
            return nusha
        else:
            return int(nusha.replace("AN", "").replace(" ", ""))

def getNameAndSurnames(columnValue):
    name=formatNameStr(columnValue.split(",")[1])
    surnames=formatNameStr(columnValue.split(",")[0]).split(" ")
    yield name
    if len(surnames)==2:
        yield surnames[0]
        yield surnames[1]
    else:
        yield columnValue.split(",")[0]
        yield ""

def getGender(columnValue):
    name=formatNameStr(columnValue.split(",")[1])
    if " " in name:
        name=name.split(" ")[0]        
    return GENDER_GUESSER_DICT[str(d.get_gender(name))]

def formatNameStr(nameStr):
    words=[]
    finalString=""
    if " " in nameStr:
        words=nameStr.split(" ")
    else:
        words.append(nameStr)
    for i in words:
        if finalString:
            finalString+=" "+i.capitalize()
        else:
            finalString+=i.capitalize()
    return finalString

def getTreatmentType(treatmentStr):
    return TREATMENT_DICT[treatmentStr]

def getTreatmentNumber(nusha):
    if nusha in patientTreatmentNumberTuple[0]:
        patientTreatmentNumberTuple[1][patientTreatmentNumberTuple[0].index(nusha)]+=1
        return patientTreatmentNumberTuple[1][patientTreatmentNumberTuple[0].index(nusha)]
    else:
        patientTreatmentNumberTuple[0].append(nusha)
        patientTreatmentNumberTuple[1].append(0)
        return 0

def getPosibilities(column):
    posibilities=[]
    for i in range(1, spreadshit.max_row):
        obj=spreadshit.cell(i, column).value
        if not(obj in posibilities):
            posibilities.append(obj)
    print(posibilities)

def showColumns():
    for i in range(1, spreadshit.max_column):
        print(str(i), ":", str(spreadshit.cell(3, i).value))

def getExtraStrings(rowNumber):
    yield spreadshit.cell(rowNumber, 48).value    #Doctor
    yield spreadshit.cell(rowNumber, 53).value    #Radiofísico
    yield spreadshit.cell(rowNumber, 27).value    #Doctors observation
    yield spreadshit.cell(rowNumber, 38).value    #Radiofísico observation

def getDates(rowNumber):
    dates=dateDefiningString(DATE_OPTIONS.index("Solicitud"), spreadshit.cell(rowNumber, 28).value.toordinal(), 0)    #Request
    dates+=","+dateDefiningString(DATE_OPTIONS.index("Recepción completa"), spreadshit.cell(rowNumber, 34).value.toordinal(), 0)
    completeReception=spreadshit.cell(rowNumber, 34).value.toordinal()
    for i in qaColumns:
        qaDate=spreadshit.cell(rowNumber, i).value
        if qaDate:
            dates+=","+dateDefiningString(DATE_OPTIONS.index("QA"), qaDate.toordinal(), completeReception)
    rawDate=spreadshit.cell(rowNumber, 35).value
    if rawDate:
        dates+=","+dateDefiningString(DATE_OPTIONS.index("Fin del calculo (Fin DC)"), rawDate.toordinal(), completeReception)
        rawDate=spreadshit.cell(rowNumber, 36).value      
        if rawDate:  
            dates+=","+dateDefiningString(DATE_OPTIONS.index("Emisión"), rawDate.toordinal(), completeReception)
    return dates

def dateDefiningString(dateType, indate, reception):
    if reception==0:
        return (str(dateType)+":"+str(indate)+":"+"0")
    else:
        if str(indate-reception)==0:
            return (str(dateType)+":"+str(indate)+":"+"0")
        else:
            return (str(dateType)+":"+str(indate)+":"+str(indate-reception))

def getCalcTries(rowNumber):
    tries=0
    for i in treatmentTriesColumns:
        content=spreadshit.cell(rowNumber, i).value
        if (content)and(type(content)==int):
            tries+=int(content)
    return tries

if __name__ == "__main__":
    d=gender.Detector(False)
    #Stablish connection to db
    sql=MySQLdb(inputLogin())
    sql.connectToDemographicDB()
    sql.connectToTreatmentDB()
    sql.loadDemographicTable()
    sql.loadTreatmentTable()
    # Define variable to load the dataframe
    excelFile=input("Archivo excel (.xlsx):")
    loadTime=time()
    print("Cargando excel...")
    dataframe = openpyxl.load_workbook(excelFile)
    print("Excel cargado en ", str(time()-loadTime), "s")
    # Define variable to read sheet
    spreadshit = dataframe.active
    errors=0
    dateColumns=[28, 34, 73, 74, 75, 35, 36]
    qaColumns=[73, 74, 75]
    treatmentTriesColumns=[39, 41, 44]
    maxRowNumber=spreadshit.max_row
    patientTreatmentNumberTuple=([],    #NUSHA
        [])                             #Treatment number
    calcTime=time()
    for rowNumber in range(4, spreadshit.max_row):
        printProgressBar(rowNumber, maxRowNumber)
        try:
            nush=getNUSHA(spreadshit.cell(rowNumber, 3).value)
            if not(nush in patientTreatmentNumberTuple[0]):
                patientData=getNameAndSurnames(spreadshit.cell(rowNumber, 4).value)
                patientTuple=(nush,
                    next(patientData),
                    next(patientData),
                    next(patientData),
                    date.today().toordinal(),
                    getGender(spreadshit.cell(rowNumber, 4).value))
            extraStrings=getExtraStrings(rowNumber)
            treatmentTuple=(nush,   #AN
                getTreatmentNumber(nush),   #Treatment number
                getDates(rowNumber),  #Dates
                getTreatmentType(spreadshit.cell(rowNumber, 17).value), #treatment number
                next(extraStrings),  #Doctor
                next(extraStrings),  #Radiophysicist
                next(extraStrings),  #Doctors observation
                next(extraStrings),  #Physician observation
                getCalcTries(rowNumber))  #Number of calc tries
            sql.saveDemographicData(patientTuple)
            sql.saveTreatmentData(treatmentTuple)
        except:
            errors+=1
        
    print("\n")
    print("Procesado en ", str(time()-calcTime), "s")
    print("No se ha podido procesar el "+str(errors/(maxRowNumber-4)*100)+"%")
    print("Tiempo total de procesado: ", str(time()-loadTime),"s")
