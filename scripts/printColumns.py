import openpyxl
from globalVars import TREATMENT_DICT
from time import time                   

def showColumns():
    for i in range(1, spreadshit.max_column):
        print(str(i), ":", str(spreadshit.cell(3, i).value))


if __name__ == "__main__":
    # Define variable to load the dataframe
    loadTime=time()
    print("Cargando excel...")
    dataframe = openpyxl.load_workbook("dosimetria.xlsx")
    print("Excel cargado en ", str(time()-loadTime), "s")
    # Define variable to read sheet
    spreadshit = dataframe.active
    showColumns()